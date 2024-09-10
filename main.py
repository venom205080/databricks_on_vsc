from fastapi import FastAPI, File, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from typing import Optional
import subprocess
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from sql_parse import generate_excel

app = FastAPI()

# Serve static files like HTML, CSS, JavaScript
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/", response_class=HTMLResponse)
async def read_root():
    return """
    <!DOCTYPE html>
    <html>
    <body>
        <h2>Upload a File</h2>
        <form action="/uploadfile/" method="post" enctype="multipart/form-data">
            <input type="file" name="file">
            <button type="submit">Test</button>
        </form>
    </body>
    </html>
    """

@app.post("/uploadfile/")
async def upload_file(file: UploadFile = File(...)):
    content = await file.read()
    string_data = content.decode('utf-8')       
    generate_excel(string_data)
    
    # Define the path for the new file
    output_file_path = f"output_{file.filename}"
    
    # Write the content to a new file
    with open(output_file_path, "wb") as file:
        file.write(content)
    
    #append "show tables" line at the end of the file
    with open(output_file_path, 'a') as file:
        file.write("""
tables_df = spark.sql("SHOW TABLES")
print("Created tables:")
for row in tables_df.collect():
    print(f"Table created: {row['tableName']}")
""")
    
    try:
        result = subprocess.run(
                ["docker", "run", "--rm", "-v", f"{os.getcwd()}:/app", "bitnami/spark:3.5.0", "spark-submit", f"/app/{output_file_path}"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
        
        pattern = re.compile(r"Table created: (\w+)")
        table_names = pattern.findall(result.stdout)    #table_names variable store names of all the table created
        
        add_new_worksheet(table_names);
        
        return { "output": result.stdout}
    except subprocess.CalledProcessError as e:
        return {"status": "failure", "message": "PySpark code failed to run.", "error": str(e)}
        

def add_new_worksheet(table_names):
    # open the Generate excel report
        file_path = f"{os.getcwd()}/output_excel.xlsx"
        wb = load_workbook(filename=file_path)        #open workbook
        # ws = wb.create_sheet("CREATED TABLES AFTER RUN", 0)     #insert a new sheet at first position
        ws = wb["CREATED TABLES BEFORE RUN"]
        ws.column_dimensions['B'].width = 30
        bold_font = Font(bold=True)  # Define a bold font style
        ws['B1'].font = bold_font
        ws['B1'] = "CREATED TABLES AFTER RUN"
        for idx, val in enumerate(table_names):
            ws[f'B{idx+2}'] = val
        
        wb.save(filename=file_path)

# To run the app, use the command: uvicorn filename:app --reload






    # [docker run --rm -v /home/my/Desktop/my-try:/app bitnami/spark:3.5.0 spark-submit /app/test.py]
    # ["docker", "run", "--rm", "-v", f"{os.path.abspath(file_path)}:/app/{filename}", "pyspark-docker-image", "spark-submit", f"/app/{filename}"],