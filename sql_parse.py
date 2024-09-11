import re
import xlsxwriter
import os
from openpyxl import load_workbook
from openpyxl.styles import Font


create_table_patterns = [
        r'CREATE\s+TABLE\s+(IF\s+NOT\s+EXISTS\s+)?([\w.]+)',  # CREATE TABLE statements
        r'CREATE\s+EXTERNAL\s+TABLE\s+(IF\s+NOT\s+EXISTS\s+)?([\w.]+)',  # CREATE TABLE statements
    ]
create_view_pattern = [
    r'CREATE\s+VIEW\s+(IF\s+NOT\s+EXISTS\s+)?([\w.]+)',  # CREATE VIEW statements
    r'CREATE\s+OR\s+REPLACE\s+TEMP\s+VIEW\s+([\w.]+)',  # CREATE VIEW statements
    r'CREATE\s+OR\s+REPLACE\s+TEMPORARY\s+VIEW\s+([\w.]+)',
]
drop_table_patterns = [
        r'DROP\s+TABLE\s+(IF\s+EXISTS\s+)?([\w.]+)',        # DROP TABLE statement
    ]
drop_view_pattern = [
    r'DROP\s+VIEW\s+(IF\s+EXISTS\s+)?([\w.]+)',        # DROP TABLE statement
]

alter_table_pattern= [
        r'ALTER\s+TABLE\s+([\w.]+)'                         # ALTER TABLE statements
    ]

pattern_array = [
    create_table_patterns,
    create_view_pattern,
    drop_table_patterns,
    drop_view_pattern,
    alter_table_pattern,
]

def extract_table_names(sql_code, patterns):
    table_names = set()
    for pattern in patterns:
        regex = re.compile(pattern, re.IGNORECASE)
        matches = regex.findall(sql_code)
        for match in matches:
            if isinstance(match, tuple):
                # The table name is in the second group of the tuple
                table_name = match[1]
            else:
                # Directly extract the table name from match
                table_name = match
            table_names.add(table_name.strip())
            
    return table_names

pattern_array_name = [
    "Created Tables",
    "Created Views",
    "Dropped Tables",
    "Dropped Views",
    "Altered tables",
]

def generate_excel(sql_code):
    wb = xlsxwriter.Workbook('output_excel.xlsx')
    cell_format = wb.add_format({'bold': True})

    for index, i in enumerate(pattern_array):
        worksheet = wb.add_worksheet(f'{pattern_array_name[index].upper()}')            #adding worksheet name
        worksheet.set_column(0, 0, 30)                          #setting column width
        worksheet.write('A1', pattern_array_name[index].upper(), cell_format)           #adding header
        for idx, val in enumerate(extract_table_names(sql_code, i)):
            worksheet.write(f'A{idx+2}', val)
        
    wb.close()
    
#function to generate worksheet created tables after run
def validated_table_worksheet(table_names):
    # open the Generate excel report
    file_path = f"{os.getcwd()}/output_excel.xlsx"
    wb = load_workbook(filename=file_path)        #open workbook
    ws = wb["CREATED TABLES"]            #select worksheet
    ws.column_dimensions['B'].width = 30
    bold_font = Font(bold=True)  # Define a bold font style
    ws['B1'].font = bold_font
    ws['B1'] = "CREATED TABLES AFTER VALIDATION"
    for idx, val in enumerate(table_names):
        ws[f'B{idx+2}'] = val
        
    wb.save(filename=file_path)
    
    
#function to add errors to the worksheet
def add_errors_to_worksheet(stdout):
    
    # Regex pattern to match .(line <number>, pos <number>)
    pattern = r'\.\(line \d+, pos \d+\)'

    # Split the result.stdout into lines
    lines = stdout.splitlines()

    # List to hold lines that match the pattern
    matching_lines = []

    # Search for the pattern in each line and add matching lines to the list
    for line in lines:
        if re.search(pattern, line):
            matching_lines.append(line)    
    
    file_path = f"{os.getcwd()}/output_excel.xlsx"
    wb = load_workbook(filename=file_path) 
    ws = wb.create_sheet("CRITICAL ERRORS")
    ws.column_dimensions['A'].width = 100
    bold_font = Font(bold=True)  # Define a bold font style
    ws['A1'].font = bold_font
    ws['A1'] = "CRITICAL ERRORS"
    for idx, val in enumerate(matching_lines):
        ws[f'A{idx+2}'] = val
    
    wb.save(filename=file_path)
    
    
    
